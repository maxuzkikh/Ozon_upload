import openpyxl
import pandas as pd
import shutil

# Define paths
template_path = r'C:\Users\Max\Documents\GitHub\Ozon_upload\создание карточки XLS\Ozon_template.xlsx'
output_file_path = r'C:\Users\Max\Documents\GitHub\Ozon_upload\создание карточки XLS\output_data_for_Ozon.xlsx'
data_path = r'C:\Users\Max\Documents\GitHub\Ozon_upload\barcode\Data path barcode.xlsx'

# Copy the original workbook to a new file
shutil.copy(template_path, output_file_path)

# Load the copied workbook
wb_output = openpyxl.load_workbook(output_file_path)

# Load the data from Data path barcode.xlsx
df = pd.read_excel(data_path)

# Access the desired sheet in the copied workbook
ws_output = wb_output['Шаблон']

def find_column(ws, header_text):
    for cell in ws[2]:
        if cell.value == header_text:
            return cell.column
    raise ValueError(f"Column '{header_text}' not found in the worksheet.")


# Fill the template with data
for i, row in enumerate(df.itertuples(), start=4):
    ws_output.cell(row=i, column=find_column(ws_output, 'Артикул*'), value=row.Артикул)
    ws_output.cell(row=i, column=find_column(ws_output, 'Цена, руб.*'), value=row.Цена)
    ws_output.cell(row=i, column=find_column(ws_output, 'НДС, %*'), value=row.НДС)

    ws_output.cell(row=i, column=find_column(ws_output, 'Ширина упаковки, мм*'), value=row.Ширина упаковки * 10)
    #ws_output.cell(row=i, column=find_column(ws_output, 'Высота упаковки, мм*'), value=2 if row._5 == 1 else row._5 * 10)
    #ws_output.cell(row=i, column=find_column(ws_output, 'Длина упаковки, мм*'), value=row._6 * 10)
    #ws_output.cell(row=i, column=find_column(ws_output, 'Ссылка на главное фото*'), value=row._7)
    #ws_output.cell(row=i, column=find_column(ws_output, 'Ссылки на дополнительные фото'), value=row._8)
    #ws_output.cell(row=i, column=find_column(ws_output, 'Бренд'), value=row._9)
    #ws_output.cell(row=i, column=find_column(ws_output, 'Название модели (для объединения в одну карточку)*'), value=row._10)
    #ws_output.cell(row=i, column=find_column(ws_output, 'Цвет товара'), value=row._11)
    #ws_output.cell(row=i, column=find_column(ws_output, 'Название цвета'), value=row._12)
    #ws_output.cell(row=i, column=find_column(ws_output, 'Тип*'), value=row._13)

# Save the modified workbook
wb_output.save(output_file_path)

