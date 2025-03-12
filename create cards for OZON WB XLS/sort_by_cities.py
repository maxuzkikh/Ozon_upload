import pandas as pd
import os
import tkinter as tk
from tkinter import filedialog, simpledialog
import xlsxwriter
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# Открываем окно выбора файла
tk.Tk().withdraw()
file_path = filedialog.askopenfilename(title="Выберите Excel-файл", filetypes=[("Excel files", "*.xlsx;*.xls")])

if not file_path:
    print("Файл не выбран!")
    exit()

# Загружаем Excel
xls = pd.ExcelFile(file_path)
sheet_name = xls.sheet_names[0]  # Берем первый лист

df = pd.read_excel(xls, sheet_name=sheet_name)

# Удаляем ненужные столбцы
columns_to_remove = ["Среднее количество заказов в день, шт", "Остатки склад ВБ, шт", "Остатки МП, шт", "дней", "сумма"]
df = df.drop(columns=[col for col in columns_to_remove if col in df.columns], errors='ignore')

# Запрашиваем у пользователя количество городов
num_cities = simpledialog.askinteger("Разделение по городам", "На сколько городов разделить напечатанные наклейки?", minvalue=1)

if num_cities is None:
    print("Ввод отменен!")
    exit()

# Запрашиваем у пользователя базовое число для диапазона
base_range = simpledialog.askinteger("Диапазон сумм", "Введите базовое число для диапазона сумм:", minvalue=1)

if base_range is None:
    print("Ввод отменен!")
    exit()

lower_bound = base_range - 16
upper_bound = base_range + 16

# Проверяем, есть ли колонка Num_Copies
if 'Num_Copies' not in df.columns:
    print("Ошибка: В файле отсутствует колонка 'Num_Copies'!")
    exit()

# Добавляем новые столбцы и распределяем значения Num_Copies
for i in range(1, num_cities + 1):
    df[f"Г{i}"] = 0  # Заполняем столбцы нулями

# Равномерно распределяем значения из Num_Copies
for index, row in df.iterrows():
    total_copies = row['Num_Copies']
    base_amount = total_copies // num_cities
    remainder = total_copies % num_cities
    
    for i in range(1, num_cities + 1):
        df.at[index, f"Г{i}"] = base_amount + (1 if i <= remainder else 0)

# Определяем путь сохранения
save_dir = os.path.dirname(file_path)
file_name = "Упаковка_" + os.path.basename(file_path)
new_file_path = os.path.join(save_dir, file_name)

# Сохраняем в новый файл
df.to_excel(new_file_path, index=False, engine='openpyxl')

# Загружаем Excel с openpyxl
wb = load_workbook(new_file_path)
sheet = wb.active

# Определяем цвета для групп
fill_colors = ["FFFF99", "99CCFF", "FFCCCC", "CCFFCC", "CCCCFF"]  # Разные цвета групп

# Окрашиваем группы по сумме значений в каждом из столбцов Г1, Г2, ...
for i in range(1, num_cities + 1):
    column_name = f"Г{i}"
    col_index = None
    
    # Ищем индекс нужного столбца
    for col in range(1, sheet.max_column + 1):
        header = sheet.cell(row=1, column=col).value
        if header == column_name:
            col_index = col
            break
    
    if col_index is None:
        raise ValueError(f"Column '{column_name}' not found in the Excel file")
    
    color_index = 0
    sum_counter = 0
    group_rows = []
    
    for row in range(2, sheet.max_row + 1):
        value = sheet.cell(row=row, column=col_index).value
        if value is not None:
            sum_counter += value
            group_rows.append(row)
        
        if lower_bound <= sum_counter <= upper_bound:
            fill_color = fill_colors[color_index % len(fill_colors)]
            for r in group_rows:
                sheet.cell(row=r, column=col_index).fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
            sum_counter = 0
            group_rows = []
            color_index += 1

wb.save(new_file_path)

print(f"Файл сохранен как {new_file_path}")
