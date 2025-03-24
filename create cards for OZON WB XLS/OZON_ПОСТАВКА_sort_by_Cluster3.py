import pandas as pd
from tkinter import Tk, filedialog
import openpyxl
from openpyxl.styles import PatternFill
import os
import numpy as np

# Открытие окна выбора файлов
Tk().withdraw()  # Скрыть основное окно
file_paths = filedialog.askopenfilenames(title="Выберите файлы Excel", filetypes=[("Excel files", "*.xlsx")])

if not file_paths:
    print("Файлы не выбраны.")
    exit()

# Чтение данных из файлов, удаление первых 4 строк
dataframes = [pd.read_excel(file, skiprows=4) for file in file_paths]

# Объединение файлов в один DataFrame
data = pd.concat(dataframes)

# Группировка по "артикул" и "кластер" с суммированием "количество"
aggregated_data = data.groupby(["артикул", "кластер"], as_index=False).agg({"количество": "sum"})

# Сортировка по "кластер" и "количество" в порядке убывания
aggregated_data = aggregated_data.sort_values(by=["кластер", "количество"], ascending=[True, False])

# Добавление колонки "место" из файла с путями
barcode_file_path = "C:/Users/Max/Documents/GitHub/Ozon_upload/barcode/Data path barcode.xlsx"
barcode_data = pd.read_excel(barcode_file_path)

# Приведение имен колонок к общему виду
barcode_data.columns = barcode_data.columns.str.strip().str.lower()
aggregated_data.columns = aggregated_data.columns.str.strip().str.lower()

# Объединение данных по артикулу
aggregated_data = aggregated_data.merge(barcode_data[["артикул", "место"]], on="артикул", how="left")

# Сохраняем промежуточный результат в файл
intermediate_file = "C:/Users/Max/Documents/GitHub/Ozon_upload/barcode/intermediate_result.xlsx"
aggregated_data.to_excel(intermediate_file, index=False)

# Читаем сохраненный результат для дальнейшей обработки
data = pd.read_excel(intermediate_file)

data.columns = data.columns.str.lower()

# Проверяем, есть ли нужные колонки в файле с баркодами
if "артикул" in barcode_data.columns and "тип упорядочить" in barcode_data.columns:
    barcode_data = barcode_data[["артикул", "тип упорядочить"]]
else:
    print("Ошибка: в файле с баркодами нет колонок 'Артикул' и 'Тип упорядочить'")
    exit()

# Создаем сводную таблицу по кластерам
pivot_table = data.pivot_table(index=["артикул"], columns="кластер", values="количество", aggfunc="sum").reset_index()

pivot_table.fillna(0, inplace=True)

for col in pivot_table.columns[1:]:
    pivot_table[col] = pivot_table[col].astype(int)

# Добавляем столбец "всего"
if "всего" not in pivot_table.columns:
    pivot_table["всего"] = pivot_table.iloc[:, 1:].sum(axis=1)

places = data.groupby("артикул")["место"].apply(lambda x: ", ".join(x.unique())).reset_index()
final_table = pivot_table.merge(places, on="артикул")
final_table = final_table.merge(barcode_data, on="артикул", how="left")
final_table.loc[:, "тип упорядочить"] = final_table["тип упорядочить"].fillna("")

# Сортируем таблицу перед расчетом
final_table = final_table.sort_values(by=["тип упорядочить", "всего"], ascending=[True, False])

# Функция округления
def round_up(value, multiple):
    return int(np.ceil(value / multiple) * multiple)

final_table["Num_Copies"] = final_table.apply(
    lambda row: round_up(row["всего"], 2) if row["тип упорядочить"].startswith("1_а4") or row["тип упорядочить"].startswith("6_а4_настройки_60")
    else round_up(row["всего"], 4) if row["тип упорядочить"].startswith("2_а5")
    else round_up(row["всего"], 2) if row["тип упорядочить"].startswith("3_термобирки")
    else row["всего"], axis=1)


final_table["осталось"] = final_table["Num_Copies"] - final_table["всего"]

# Распределение оставшихся значений
city_columns = [col for col in pivot_table.columns[1:]]
for index, row in final_table.iterrows():
    remaining = row["осталось"]
    target_cities = sorted(city_columns, key=lambda city: row[city])
    if remaining > 0:
        allocation_per_city = remaining // len(target_cities)
        extra_allocation = remaining % len(target_cities)
        for city in target_cities:
            final_table.at[index, city] += allocation_per_city
        for i in range(extra_allocation):
            final_table.at[index, target_cities[i]] += 1
        final_table.at[index, "осталось"] = 0

columns_order = ["артикул"] + city_columns + ["Num_Copies", "осталось", "место", "тип упорядочить"]
final_table = final_table[columns_order]

# Функция для выделения цветом
wb = openpyxl.Workbook()
sheet = wb.active
sheet.append(final_table.columns.tolist())  # Добавляем заголовки
for row in final_table.itertuples(index=False, name=None):
    sheet.append(row)

def group_and_colorize(sheet, df, column='Num_Copies', min_sum=330, max_sum=360):
    grouped_indices = []
    current_group = []
    current_sum = 0
    colors = ['FFFF99', '99FF99', '99CCFF', 'FF9999', 'FFCC99']
    for i, value in enumerate(df[column]):
        if current_sum + value > max_sum:
            if min_sum <= current_sum <= max_sum:
                grouped_indices.append(current_group[:])
            current_group = []
            current_sum = 0
        current_group.append(i)
        current_sum += value
    if min_sum <= current_sum <= max_sum:
        grouped_indices.append(current_group)
    for i, group in enumerate(grouped_indices):
        fill = PatternFill(start_color=colors[i % len(colors)], end_color=colors[i % len(colors)], fill_type="solid")
        for row_idx in group:
            excel_row = row_idx + 2
            for col_idx in range(1, sheet.max_column + 1):
                sheet.cell(row=excel_row, column=col_idx).fill = fill

group_and_colorize(sheet, final_table)

output_path = filedialog.asksaveasfilename(title="Сохранить файл как", defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
if output_path:
    wb.save(output_path)
    print(f"Файл сохранен: {output_path}")
else:
    print("Сохранение отменено.")
