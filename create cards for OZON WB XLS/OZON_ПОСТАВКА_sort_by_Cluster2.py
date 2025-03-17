import pandas as pd
from tkinter import Tk, filedialog
import openpyxl
from openpyxl.styles import PatternFill

# Открытие окна выбора файлов
Tk().withdraw()  # Скрыть основное окно
file_paths = filedialog.askopenfilenames(title="Выберите файлы Excel", filetypes=[("Excel files", "*.xlsx")])

if not file_paths:
    print("Файлы не выбраны.")
    exit()

# Чтение данных из файлов, удаление первых 4 строк
dataframes = [pd.read_excel(file, skiprows=4) for file in file_paths]

# Объединение файлов в один DataFrame
data = pd.concat(dataframes)

# Группировка по "артикул" и "кластер" с суммированием "количество"
aggregated_data = data.groupby(["артикул", "кластер"], as_index=False).agg({"количество": "sum"})

# Сортировка по "кластер" и "количество" в порядке убывания
aggregated_data = aggregated_data.sort_values(by=["кластер", "количество"], ascending=[True, False])

# Добавление колонки "место" из файла с путями
barcode_file_path = "C:/Users/Max/Documents/GitHub/Ozon_upload/barcode/Data path barcode.xlsx"
barcode_data = pd.read_excel(barcode_file_path)

# Приведение имен колонок к общему виду
barcode_data.columns = barcode_data.columns.str.strip().str.lower()
aggregated_data.columns = aggregated_data.columns.str.strip().str.lower()

# Объединение данных по артикулу
aggregated_data = aggregated_data.merge(barcode_data[["артикул", "место"]], on="артикул", how="left")

# Выбор пути для сохранения файла
output_path = filedialog.asksaveasfilename(title="Сохранить файл как", defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
if output_path:
    # Сохранение в Excel
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        aggregated_data.to_excel(writer, index=False, sheet_name="Data")
        workbook = writer.book
        worksheet = writer.sheets["Data"]
        
        # Цвета для кластеров
        cluster_colors = {}
        colors = ["FFFFCC", "CCFFCC", "CCCCFF", "FFCCCC", "CCE5FF", "FFD9CC"]  # Разные пастельные цвета
        color_index = 0
        
        for row in range(2, worksheet.max_row + 1):
            cluster_value = worksheet.cell(row=row, column=2).value  # Колонка "кластер"
            if cluster_value not in cluster_colors:
                cluster_colors[cluster_value] = colors[color_index % len(colors)]
                color_index += 1
            
            fill = PatternFill(start_color=cluster_colors[cluster_value], end_color=cluster_colors[cluster_value], fill_type="solid")
            for col in range(1, worksheet.max_column + 1):
                worksheet.cell(row=row, column=col).fill = fill
    
    print(f"Файл сохранен: {output_path}")
else:
    print("Сохранение отменено.")