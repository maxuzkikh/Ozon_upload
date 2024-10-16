import pandas as pd
import openpyxl
import shutil
import tkinter as tk
from tkinter import filedialog
import os

# Create a Tkinter root window (it won't be shown)
root = tk.Tk()
root.withdraw()  # Hide the root window

# Prompt the user to select the Fill_data.xlsx file
data_path = filedialog.askopenfilename(title="Select Fill_data.xlsx", filetypes=[("Excel files", "*.xlsx")])

if not data_path:  # Check if the user canceled the file selection
    print("No file selected, exiting...")
    exit()

# Paths to the template file
template_path = 'C:/Users/Max/Documents/GitHub/Ozon_upload/create cards for OZON WB XLS/templates/WB_template.xlsx'

# Temporary path for the filled template
temp_filled_path = 'C:/Users/Max/Documents/GitHub/Ozon_upload/create cards for OZON WB XLS/templates/temp_WB_template_upload.xlsx'

# Copy the template file to a temporary filled path
shutil.copyfile(template_path, temp_filled_path)
print(f"Template successfully copied to: {temp_filled_path}")

# Load the data from the selected Fill_data.xlsx file
data_df = pd.read_excel(data_path, sheet_name=0, header=0)  # Headers in data start from the first row

# Open the copied template file using openpyxl
wb = openpyxl.load_workbook(temp_filled_path)
ws = wb.active  # Active sheet

# Mapping the column headers
columns_map = {
    'Артикул продавца': 'Артикул продавца',
    'Наименование': 'Наименование',
    'Категория продавца': 'Категория продавца',
    'Бренд': 'Бренд',
    'Описание': 'Описание',
    'Фото': 'Фото',
    'Видео': 'Видео',
    'Баркоды': 'Баркоды',
    'Высота упаковки': 'Высота упаковки',
    'Длина упаковки': 'Длина упаковки',
    'Вес товара с упаковкой (г)': 'Вес товара с упаковкой (г)', 
    'Цена': 'Цена'
}

# Find the columns in the template based on their headers
template_columns = {cell.value: idx + 1 for idx, cell in enumerate(ws[3])}  # Assuming headers are on the 3rd row

# Check for necessary headers in the data file
for data_col in columns_map.keys():
    if data_col not in data_df.columns:
        print(f"Warning! Column '{data_col}' not found in the data and will be skipped.")

# Fill only the necessary columns with data, starting from row 5
for i, row in data_df.iterrows():
    for data_col, template_col in columns_map.items():
        if data_col in data_df.columns and template_col in template_columns:  # Check if the column exists
            ws.cell(row=i + 5, column=template_columns[template_col], value=row[data_col])

# Extract the base name from the selected file
base_name = os.path.splitext(os.path.basename(data_path))[0]  # Get the filename without extension
final_filename = f"{base_name}_data_for_WB.xlsx"  # Create new filename with suffix

# Prompt the user to choose where to save the final filled file
save_path = filedialog.asksaveasfilename(defaultextension=".xlsx",
                                           title="Save as",
                                           filetypes=[("Excel files", "*.xlsx")],
                                           initialfile=final_filename)

if not save_path:  # Check if the user canceled the save dialog
    print("No save location selected, exiting...")
    exit()

# Save the filled template to the chosen location
wb.save(save_path)
print(f"Final file successfully filled and saved to: {save_path}")

# Optionally, remove the temporary filled file
shutil.os.remove(temp_filled_path)
