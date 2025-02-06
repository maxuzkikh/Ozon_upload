import pandas as pd
import tkinter as tk
from tkinter import filedialog
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# Открыть окно выбора файла
def select_files():
    root = tk.Tk()
    root.withdraw()  # Скрыть главное окно
    file_paths = filedialog.askopenfilenames(filetypes=[("Excel files", "*.xlsx *.xls")])
    return file_paths

# Основной скрипт
file_paths = select_files()

if len(file_paths) == 2:
    try:
        # Чтение первого и второго файлов
        file1_path, file2_path = file_paths
        df1 = pd.read_excel(file1_path, header=None)
        df2 = pd.read_excel(file2_path, header=None)

        # Убедимся, что оба файла содержат достаточное количество данных
        if df1.shape[0] < 6 or df2.shape[0] < 6:
            raise ValueError("В одном из файлов недостаточно данных.")

        # Из второго файла берем строки начиная с А6
        df2_data = df2.iloc[5:].reset_index(drop=True)

        # Добавляем строки из второго файла в конец первого
        df1 = pd.concat([df1, df2_data], ignore_index=True)

        # Продолжаем логику с обработкой первого файла
        # Сортируем строки начиная с 6-й
        df_header = df1.iloc[:5]  # Сохраняем заголовки
        df_data = df1.iloc[5:]  # Данные для сортировки
        df_data_sorted = df_data.sort_values(by=4)  # Сортировка по 5-му столбцу ("кластер")

        # Объединяем заголовки и отсортированные данные
        df1 = pd.concat([df_header, df_data_sorted], ignore_index=True)

        # Удаляем первые 4 строки
        df1 = df1.iloc[4:].reset_index(drop=True)

        # Удаляем ненужные столбцы
        columns_to_drop = [0, 2] + list(range(5, 15))
        df1 = df1.drop(columns=[col for col in columns_to_drop if col in df1.columns], axis=1)

        # Переименование столбца и добавление нового
        df1.rename(columns={df1.columns[0]: "Артикул"}, inplace=True)
        df1.insert(3, "место", "")  # Вставляем столбец "место" в 4-й индекс (D)

        # Чтение данных из второго файла для добавления "место"
        df2_additional = pd.read_excel(r"C:\Users\Max\Documents\GitHub\Ozon_upload\barcode\Data path barcode.xlsx")

        # Проверяем наличие нужных столбцов
        if "Артикул" not in df2_additional.columns or "место" not in df2_additional.columns:
            raise ValueError("Отсутствуют необходимые столбцы в файле данных.")

        # Объединяем данные по столбцу "Артикул"
        df1 = df1.merge(df2_additional[["Артикул", "место"]], on="Артикул", how="left")

        # Сохраняем результат в новый файл
        save_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx *.xls")])
        if save_path:
            df1.to_excel(save_path, index=False, header=True)

            # Устанавливаем ширину столбца A
            wb = load_workbook(save_path)
            ws = wb.active
            ws.column_dimensions[get_column_letter(1)].width = 40  # Примерная ширина
            wb.save(save_path)

            print(f"Файл успешно сохранён: {save_path}")
        else:
            print("Сохранение отменено.")
    except Exception as e:
        print(f"Ошибка: {e}")
else:
    print("Необходимо выбрать ровно два файла.")
