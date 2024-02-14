import openpyxl
import pandas as pd
import shutil

# Define paths
template_path = r'C:\Users\Max\Documents\GitHub\Ozon_upload\create cards for OZON WB XLS\Ozon_template.xlsx'
output_file_path = r'C:\Users\Max\Documents\GitHub\Ozon_upload\create cards for OZON WB XLS\output_data_for_Ozon.xlsx'
data_path = r'C:\Users\Max\Documents\GitHub\Ozon_upload\create cards for OZON WB XLS\Data_to_create.xlsx'

# Copy the original workbook to a new file
shutil.copy(template_path, output_file_path)

# Load the copied workbook
wb_output = openpyxl.load_workbook(output_file_path)

# Load the data from Data path barcode.xlsx
df = pd.read_excel(data_path)

# Access the desired sheet in the copied workbook
ws_output = wb_output['Шаблон']
ws_output2 = wb_output['Озон.Видео']
# Get all column names
column_names = [cell.value for cell in ws_output[2]]


def find_column(workbook_path, sheet_name, header_text):
    wb = openpyxl.load_workbook(workbook_path)
    ws = wb[sheet_name]
    #print(f"Searching for '{header_text}' in '{sheet_name}'")
    if workbook_path == data_path:
        for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=1), start=1):
            for cell in row:
                if cell.value == header_text:
                    return cell.column, row_idx
        raise ValueError(f"Column '{header_text}' not found in the worksheet '{sheet_name}'.")
    if workbook_path == output_file_path:
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=2), start=2):
            for cell in row:
                if cell.value == header_text:
                    return cell.column, row_idx
        raise ValueError(f"Column '{header_text}' not found in the worksheet '{sheet_name}'.")

# Find column indices outside the loop
article_col, _ = find_column(data_path, 'Лист1', 'Артикул')
poduct_name_col, _ = find_column(data_path, 'Лист1', 'Наименование')
price_col, _ = find_column(data_path, 'Лист1', 'Цена')
vat_col, _ = find_column(data_path, 'Лист1', 'НДС, %*')
promote_col, _ = find_column(data_path, 'Лист1', 'Включить продвижение')
weight_col, _ = find_column(data_path, 'Лист1', 'Вес товара с упаковкой (г)')
width_col, _ = find_column(data_path, 'Лист1', 'Ширина упаковки, мм*')
height_col, _ = find_column(data_path, 'Лист1', 'Высота упаковки, мм*')
lenght_col, _ = find_column(data_path, 'Лист1', 'Длина упаковки, мм*')
photo_main_col, _ = find_column(data_path, 'Лист1', 'Ссылка на главное фото*')
photo_aditional_col, _ = find_column(data_path, 'Лист1', 'Ссылки на дополнительные фото')
brand_col, _ = find_column(data_path, 'Лист1', 'Бренд')
combine_col, _ = find_column(data_path, 'Лист1', 'Название модели (для объединения в одну карточку)*')
color_name_col, _ = find_column(data_path, 'Лист1', 'Цвет товара')
type_col, _ = find_column(data_path, 'Лист1', 'Тип*')
annotation_col, _ = find_column(data_path, 'Лист1', 'Аннотация')
country_col, _ = find_column(data_path, 'Лист1', 'Страна производства')
material_col, _ = find_column(data_path, 'Лист1', 'Состав')
color_example_col, _ = find_column(data_path, 'Лист1', 'Образец цвета')
key_words_col, _ = find_column(data_path, 'Лист1', 'Ключевые слова')
video_name_col, _ = find_column(data_path, 'Лист1', 'Озон.Видео: название')
video_link_col, _ = find_column(data_path, 'Лист1', 'Озон.Видео: ссылка')

print(article_col,price_col,vat_col,weight_col,width_col,height_col,lenght_col,photo_main_col,photo_aditional_col)

print(f"Please Wait when its DONE....")
# Fill the template with data
for i, row in enumerate(df.itertuples(), start=4):
    ws_output.cell(row=i, column=find_column(output_file_path, 'Шаблон', 'Артикул*')[0], value=row[article_col])
    ws_output.cell(row=i, column=find_column(output_file_path, 'Шаблон', 'Название товара')[0], value=row[poduct_name_col])
    ws_output.cell(row=i, column=find_column(output_file_path, 'Шаблон', 'Цена, руб.*')[0], value=row[price_col])
    ws_output.cell(row=i, column=find_column(output_file_path, 'Шаблон', 'НДС, %*')[0], value=row[vat_col])
    ws_output.cell(row=i, column=find_column(output_file_path, 'Шаблон', 'Включить продвижение')[0], value=row[promote_col])
    ws_output.cell(row=i, column=find_column(output_file_path, 'Шаблон', 'Вес в упаковке, г*')[0], value=row[weight_col])
    ws_output.cell(row=i, column=find_column(output_file_path, 'Шаблон', 'Ширина упаковки, мм*')[0], value=row[width_col])
    ws_output.cell(row=i, column=find_column(output_file_path, 'Шаблон', 'Высота упаковки, мм*')[0], value=row[height_col])
    ws_output.cell(row=i, column=find_column(output_file_path, 'Шаблон', 'Длина упаковки, мм*')[0], value=row[lenght_col])
    ws_output.cell(row=i, column=find_column(output_file_path, 'Шаблон', 'Ссылка на главное фото*')[0], value=row[photo_main_col])
    ws_output.cell(row=i, column=find_column(output_file_path, 'Шаблон', 'Ссылки на дополнительные фото')[0], value=row[photo_aditional_col])
    ws_output.cell(row=i, column=find_column(output_file_path, 'Шаблон', 'Бренд*')[0], value=row[brand_col])
    ws_output.cell(row=i, column=find_column(output_file_path, 'Шаблон', 'Название модели (для объединения в одну карточку)*')[0], value=row[combine_col])
    ws_output.cell(row=i, column=find_column(output_file_path, 'Шаблон', 'Название цвета')[0], value=row[color_name_col])
    ws_output.cell(row=i, column=find_column(output_file_path, 'Шаблон', 'Тип*')[0], value=row[type_col])
    ws_output.cell(row=i, column=find_column(output_file_path, 'Шаблон', 'Аннотация')[0], value=row[annotation_col])
    ws_output.cell(row=i, column=find_column(output_file_path, 'Шаблон', 'Страна-изготовитель')[0], value=row[country_col])
    ws_output.cell(row=i, column=find_column(output_file_path, 'Шаблон', 'Материал')[0], value=row[material_col])
    ws_output.cell(row=i, column=find_column(output_file_path, 'Шаблон', 'Образец цвета')[0], value=row[color_example_col])
    ws_output.cell(row=i, column=find_column(output_file_path, 'Шаблон', 'Ключевые слова')[0], value=row[key_words_col])
    ws_output2.cell(row=i, column=find_column(output_file_path, 'Озон.Видео', 'Озон.Видео: название')[0], value=row[video_name_col])
    ws_output2.cell(row=i, column=find_column(output_file_path, 'Озон.Видео', 'Озон.Видео: ссылка')[0], value=row[video_link_col])
    ws_output2.cell(row=i, column=find_column(output_file_path, 'Озон.Видео', 'Артикул*')[0], value=row[article_col])

# Save the modified workbook
wb_output.save(output_file_path)
print("Done")