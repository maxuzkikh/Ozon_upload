import os
import math
import io
import pandas as pd
from tkinter import Tk, filedialog
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from PyPDF2 import PdfMerger, PdfReader, PdfWriter

# Выбор Excel-файла
Tk().withdraw()
excel_file_path = filedialog.askopenfilename(title="Выберите Excel-файл", filetypes=[("Excel files", "*.xlsx")])
if not excel_file_path:
    raise Exception("Файл не выбран!")

output_dir = os.path.dirname(excel_file_path)

# Пути к данным
barcode_path = r"C:\Users\Max\Documents\GitHub\Ozon_upload\barcode\Data path barcode.xlsx"

# Чтение файлов
df_main = pd.read_excel(excel_file_path)
df_barcode = pd.read_excel(barcode_path)

if "Тип упорядочить" not in df_barcode.columns:
    raise KeyError("Не найдена колонка 'Тип упорядочить' в barcode-файле")

# Объединение и сортировка
merged_df = pd.merge(df_main, df_barcode, on="Артикул")
merged_df = merged_df.sort_values(by=["Тип упорядочить", "Num_Copies"], ascending=[True, False])
merged_df = merged_df[merged_df['Num_Copies'] != 0]

# Корректировка Num_Copies
def adjust_num_copies(row):
    if row["Тип упорядочить"] in ["1_а4", "6_а4_настройки_60", "3_термобирки"]:
        return math.ceil(row["Num_Copies"] / 2) * 2
    elif row["Тип упорядочить"] == "2_а5":
        return math.ceil(row["Num_Copies"] / 4) * 4
    return row["Num_Copies"]

merged_df["Num_Copies"] = merged_df.apply(adjust_num_copies, axis=1)
columns_to_keep = list(df_main.columns) + ["Тип упорядочить"]
columns_to_keep = [col for col in columns_to_keep if col in merged_df.columns]
sorted_path = excel_file_path.replace('.xlsx', '_sorted.xlsx')
merged_df[columns_to_keep].to_excel(sorted_path, index=False, engine='openpyxl')

# Окрашивание Excel + Формирование групп
wb = load_workbook(sorted_path)
sheet = wb.active
fill_colors = ["FFFF99", "99CCFF", "FFCCCC", "CCFFCC", "CCCCFF"]
special_fill = "FFA07A"

# Найти номера колонок
num_col = None
type_col = None
for col in range(1, sheet.max_column + 1):
    val = sheet.cell(row=1, column=col).value
    if val == "Num_Copies":
        num_col = col
    elif val == "Тип упорядочить":
        type_col = col

MAX_GROUP = 350
sum_counter = 0
group_rows = []
color_index = 0
special_rows = []
all_groups = []

for row in range(2, sheet.max_row + 1):
    copies = sheet.cell(row=row, column=num_col).value
    group_type = sheet.cell(row=row, column=type_col).value

    if group_type == "6_а4_настройки_60":
        special_rows.append(row)
        continue

    if copies is not None:
        if sum_counter + copies > MAX_GROUP:
            color = fill_colors[color_index]
            for r in group_rows:
                for c in range(1, sheet.max_column + 1):
                    sheet.cell(row=r, column=c).fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
            all_groups.append(group_rows[:])
            group_rows = []
            sum_counter = 0
            color_index = (color_index + 1) % len(fill_colors)

        sum_counter += copies
        group_rows.append(row)

# Последняя группа
if group_rows:
    color = fill_colors[color_index]
    for r in group_rows:
        for c in range(1, sheet.max_column + 1):
            sheet.cell(row=r, column=c).fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
    all_groups.append(group_rows[:])

# Спец-группа
for r in special_rows:
    for c in range(1, sheet.max_column + 1):
        sheet.cell(row=r, column=c).fill = PatternFill(start_color=special_fill, end_color=special_fill, fill_type="solid")

wb.save(sorted_path)

# Чтение окрашенного файла
df_sorted = pd.read_excel(sorted_path, engine='openpyxl')

# Экспорт групп
for idx, group in enumerate(all_groups, start=1):
    df_group = df_sorted.iloc[[r - 2 for r in group]]  # Excel → DataFrame index
    p_file = os.path.join(output_dir, f"p{idx}.xlsx")
    df_group.to_excel(p_file, index=False, engine='openpyxl')
    print(f"✅ Группа {idx} экспортирована: {p_file}")

    # Объединение PDF
    df_paths = pd.read_excel(barcode_path)
    df_merged = pd.merge(df_group, df_paths[['Артикул', 'Local_PDF_Path_Column_Name']], on='Артикул', how='left')

    missing = df_merged[df_merged['Local_PDF_Path_Column_Name'].isna()]
    if not missing.empty:
        print(f"⚠️ Нет пути к PDF для артикулов: {missing['Артикул'].tolist()}")

    merger = PdfMerger()
    for _, row in df_merged.iterrows():
        pdf_path = row['Local_PDF_Path_Column_Name']
        num_copies = int(row['Num_Copies'])

        if not os.path.isfile(pdf_path):
            print(f"❌ Файл не найден: {pdf_path}")
            continue

        try:
            reader = PdfReader(pdf_path)
            if len(reader.pages) == 0:
                print(f"⚠️ Пустой PDF: {pdf_path}")
                continue
            first_page = reader.pages[0]

            for _ in range(num_copies):
                writer = PdfWriter()
                writer.add_page(first_page)
                buffer = io.BytesIO()
                writer.write(buffer)
                buffer.seek(0)
                merger.append(buffer)

        except Exception as e:
            print(f"‼️ Ошибка при обработке {pdf_path}: {e}")

    final_pdf_path = os.path.join(output_dir, f"{os.path.splitext(os.path.basename(p_file))[0]}_печать.pdf")
    merger.write(final_pdf_path)
    merger.close()
    print(f"📄 PDF для группы {idx} сохранён: {final_pdf_path}")

print("🎉 Все PDF-файлы сгенерированы!")
