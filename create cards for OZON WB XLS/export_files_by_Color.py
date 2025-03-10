import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import os
from tkinter import Tk
from tkinter.filedialog import askopenfilename

# Открываем окно выбора файла
Tk().withdraw()
sorted_excel_file_path = askopenfilename(title="Выберите обработанный Excel файл", filetypes=[("Excel files", "*.xlsx")])
output_dir = "exported_groups"  # Директория для сохранения файлов

# Создадим папку для экспорта, если её нет
os.makedirs(output_dir, exist_ok=True)

# Загрузим обработанный Excel
wb = load_workbook(sorted_excel_file_path)
sheet = wb.active

def get_color(cell):
    """Функция для получения цвета ячейки"""
    fill = cell.fill
    return fill.start_color.rgb if fill.fill_type == "solid" else None

# Определим цвета и строки
color_groups = {}
for row in range(2, sheet.max_row + 1):
    color = get_color(sheet.cell(row=row, column=1))  # Берем цвет из первой колонки
    if color:
        if color not in color_groups:
            color_groups[color] = []
        color_groups[color].append(row)

# Читаем оригинальный DataFrame
df_sorted = pd.read_excel(sorted_excel_file_path, engine='openpyxl')

# Экспортируем группы по цвету в отдельные файлы
for idx, (color, rows) in enumerate(color_groups.items(), start=1):
    df_group = df_sorted.iloc[[r - 2 for r in rows]]  # Смещение на 2, так как первая строка — заголовок
    file_path = os.path.join(output_dir, f"p{idx}.xlsx")
    df_group.to_excel(file_path, index=False, engine='openpyxl')
    print(f"Группа {idx} ({color}) экспортирована в {file_path}")

print("Все цветовые группы экспортированы!")