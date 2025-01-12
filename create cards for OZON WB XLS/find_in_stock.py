import shutil
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Border, Side
from tkinter import Tk
from tkinter.filedialog import askopenfilename
import os
import win32print
import win32api
from tkinter import messagebox, Tk
import time

def select_excel_file():
    Tk().withdraw()  # Hide the root window
    filename = askopenfilename(filetypes=[("Excel files", "*.xlsx")])
    return filename

def copy_and_modify_excel(file_path):
    folder_path = os.path.dirname(file_path)
    file_name = os.path.basename(file_path)
    new_file_path = os.path.join(folder_path, f"copy_{file_name}")
    shutil.copy(file_path, new_file_path)

    wb = load_workbook(new_file_path)
    ws = wb.active

    ws['I5'] = 'место'
    ws.column_dimensions['I'].width = 10
    ws.column_dimensions['G'].width = 20
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 7

    bold_font = Font(bold=True)
    ws['G5'].font = bold_font
    ws['I5'].font = bold_font

    wb.save(new_file_path)
    return new_file_path

def find_and_fill_values(file_path, barcode_file_path):
    bold_font = Font(bold=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    wb = load_workbook(file_path)
    ws = wb.active
    df = pd.read_excel(barcode_file_path)

    row = 6
    while True:
        article_value = ws[f'G{row}'].value
        if not article_value:
            break

        match_row = df[df['Артикул'] == article_value]
        if not match_row.empty:
            место_value = match_row.iloc[0]['место']
            ws[f'I{row}'] = место_value
        else:
            ws[f'I{row}'] = 'Not Found'

        ws.row_dimensions[row].height = 45
        ws[f'G{row}'].font = bold_font
        ws[f'I{row}'].font = bold_font
        ws[f'I{row}'].border = thin_border

        row += 1

    row = 6
    while True:
        cell_value = ws[f'H{row}'].value
        if not cell_value:
            break

        ws[f'H{row}'].value = cell_value[8:]
        row += 1

    ws.delete_cols(3)
    ws.delete_cols(3)
    ws.delete_cols(4)

    for row in ws.iter_rows(min_row=6, max_row=ws.max_row, min_col=5, max_col=6):
        for cell in row:
            if cell.value is not None:
                cell.font = Font(size=14, bold=True)

    wb.save(file_path)
    print(f"Values for 'место' filled and columns modified successfully in {file_path}")

def print_file(file_path):
    # Define printer name
    printer_name = "Kyocera ECOSYS P2040dn KX"
    # Ensure the printer is available
    printers = [printer[2] for printer in win32print.EnumPrinters(2)]
    if printer_name not in printers:
        print(f"Error: Printer '{printer_name}' not found.")
        return
    time.sleep(1)
    # Open a handle to the printer
    hPrinter = win32print.OpenPrinter(printer_name)
    time.sleep(1)
    # Configure printer settings (landscape or portrait orientation, A4, etc.)
    pdc = win32print.GetPrinter(hPrinter, 2)  # Level 2 info structure
    pdc["pDevMode"].PaperSize = 9  # A4
    pdc["pDevMode"].Orientation = 1  # Portrait (1) or Landscape (2)
    time.sleep(1)
    # Print the file
    try:
        win32api.ShellExecute(
            0,
            "printto",
            file_path,
            f'"{printer_name}"',
            ".",
            0
        )
        print(f"File '{file_path}' sent to printer '{printer_name}' successfully.")
    except Exception as e:
        print(f"Error during printing: {e}")
    finally:
        win32print.ClosePrinter(hPrinter)


if __name__ == "__main__":
    excel_file = select_excel_file()
    copied_excel_file = copy_and_modify_excel(excel_file)
    barcode_file_path = r"C:\Users\Max\Documents\GitHub\Ozon_upload\barcode\Data path barcode.xlsx"
    find_and_fill_values(copied_excel_file, barcode_file_path)
    
    # Define printer name
    printer_name = "Kyocera ECOSYS P2040dn"

    # Create a hidden Tkinter root window
    root = Tk()
    root.withdraw()  # Hide the main window

    # Ask the user if they want to print
    user_response = messagebox.askyesno(
        "Print Confirmation",
        f"Хотите распечатать модифицированный файл на принтере {printer_name}?"
    )

    if user_response:  # If user clicks 'Yes'
        print_file(copied_excel_file)
    else:
        print("Printing skipped.")

    # Destroy the root window
    root.destroy()

