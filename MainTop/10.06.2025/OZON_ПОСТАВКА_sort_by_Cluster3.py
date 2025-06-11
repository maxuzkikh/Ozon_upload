import pandas as pd
from tkinter import Tk, filedialog
import openpyxl
from openpyxl.styles import PatternFill
import os
import numpy as np

# Открытие окна выбора файлов
Tk().withdraw()  # Скрыть основное окно
file_paths = filedialog.askopenfilenames(title="Выберите файлы Excel", filetypes=[("Excel files", "*.xlsx")])

if not file_paths:
    print("Файлы не выбраны.")
    exit()

# Чтение данных из файлов, удаление первых 4 строк
dataframes = []
for file in file_paths:
    df = pd.read_excel(file, skiprows=5)
    df = df.rename(columns={
        df.columns[1]: "Артикул",
        df.columns[5]: "Кластер",
        df.columns[3]: "Рекомендуемая поставка, шт на 56 дней"
    })
    dataframes.append(df)



# Объединение файлов в один DataFrame
data = pd.concat(dataframes)
print(data.columns.tolist())

# Группировка по "артикул" и "кластер" с суммированием "количество"
aggregated_data = data.groupby(["Артикул", "Кластер"], as_index=False).agg({"Рекомендуемая поставка, шт на 56 дней": "sum"})

# Сортировка по "кластер" и "количество" в порядке убывания
aggregated_data = aggregated_data.sort_values(by=["Кластер", "Рекомендуемая поставка, шт на 56 дней"], ascending=[True, False])

# Добавление колонки "Место" из файла с путями
barcode_file_path = "C:/Users/maxim/Documents/GitHub/Ozon_upload/barcode/Data path barcode.xlsx"
barcode_data = pd.read_excel(barcode_file_path)
barcode_data.columns = barcode_data.columns.str.strip()  # Удаляем пробелы
print("Колонки barcode_data:", barcode_data.columns.tolist())  # Отладка

aggregated_data = aggregated_data.merge(barcode_data[["Артикул", "место"]], on="Артикул", how="left")


# Приведение имен колонок к общему виду
#barcode_data.columns = barcode_data.columns.str.strip().str.lower()
#aggregated_data.columns = aggregated_data.columns.str.strip().str.lower()



# Сохраняем промежуточный результат в файл
intermediate_file = "C:/Users/maxim/Documents/GitHub/Ozon_upload/barcode/intermediate_result.xlsx"
aggregated_data.to_excel(intermediate_file, index=False)

# Читаем сохраненный результат для дальнейшей обработки
data = pd.read_excel(intermediate_file)

data.columns = [  # Приводим названия обратно к ожидаемым
    "Артикул", "Кластер", "Рекомендуемая поставка, шт на 56 дней", "Место"
]


# Проверяем, есть ли нужные колонки в файле с баркодами
if "Артикул" in barcode_data.columns and "Тип упорядочить" in barcode_data.columns:
    barcode_data = barcode_data[["Артикул", "Тип упорядочить"]]
else:
    print("Ошибка: в файле с баркодами нет колонок 'Артикул' и 'Тип упорядочить'")
    exit()


# Создаем сводную таблицу по кластерам
pivot_table = data.pivot_table(index=["Артикул"], columns="Кластер", values="Рекомендуемая поставка, шт на 56 дней", aggfunc="sum").reset_index()

pivot_table.fillna(0, inplace=True)

for col in pivot_table.columns[1:]:
    pivot_table[col] = pivot_table[col].astype(int)

# Добавляем столбец "всего"
if "всего" not in pivot_table.columns:
    pivot_table["всего"] = pivot_table.iloc[:, 1:].sum(axis=1)

places = data.groupby("Артикул")["Место"].apply(lambda x: ", ".join(x.unique())).reset_index()
final_table = pivot_table.merge(places, on="Артикул")
final_table = final_table.merge(barcode_data, on="Артикул", how="left")
final_table.loc[:, "Тип упорядочить"] = final_table["Тип упорядочить"].fillna("")


# Сортируем таблицу перед расчетом
final_table = final_table.sort_values(by=["Тип упорядочить", "всего"], ascending=[True, False])


# Функция округления
def round_up(value, multiple):
    return int(np.ceil(value / multiple) * multiple)

final_table["Num_Copies"] = final_table.apply(
    lambda row: round_up(row["всего"], 2) if row["Тип упорядочить"].startswith("1_а4") or row["Тип упорядочить"].startswith("6_а4_настройки_60")
    else round_up(row["всего"], 4) if row["Тип упорядочить"].startswith("2_а5")
    else round_up(row["всего"], 2) if row["Тип упорядочить"].startswith("3_термобирки")
    else row["всего"], axis=1)


final_table["осталось"] = final_table["Num_Copies"] - final_table["всего"]

# Распределение оставшихся значений
city_columns = [col for col in pivot_table.columns[1:]]
for index, row in final_table.iterrows():
    remaining = row["осталось"]
    target_cities = sorted(city_columns, key=lambda city: row[city])
    if remaining > 0:
        allocation_per_city = remaining // len(target_cities)
        extra_allocation = remaining % len(target_cities)
        for city in target_cities:
            final_table.at[index, city] += allocation_per_city
        for i in range(extra_allocation):
            final_table.at[index, target_cities[i]] += 1
        final_table.at[index, "осталось"] = 0

columns_order = ["Артикул"] + city_columns + ["Num_Copies", "осталось", "Место", "Тип упорядочить"]

final_table = final_table[columns_order]

# Функция для выделения цветом
wb = openpyxl.Workbook()
sheet = wb.active
sheet.append(final_table.columns.tolist())  # Добавляем заголовки
for row in final_table.itertuples(index=False, name=None):
    sheet.append(row)

def group_and_colorize(sheet, df, column='Num_Copies', min_sum=330, max_sum=360):
    grouped_indices = []
    current_group = []
    current_sum = 0
    colors = ['FFFF99', '99FF99', '99CCFF', 'FF9999', 'FFCC99']
    for i, value in enumerate(df[column]):
        if current_sum + value > max_sum:
            if min_sum <= current_sum <= max_sum:
                grouped_indices.append(current_group[:])
            current_group = []
            current_sum = 0
        current_group.append(i)
        current_sum += value
    if min_sum <= current_sum <= max_sum:
        grouped_indices.append(current_group)
    for i, group in enumerate(grouped_indices):
        fill = PatternFill(start_color=colors[i % len(colors)], end_color=colors[i % len(colors)], fill_type="solid")
        for row_idx in group:
            excel_row = row_idx + 2
            for col_idx in range(1, sheet.max_column + 1):
                sheet.cell(row=excel_row, column=col_idx).fill = fill

group_and_colorize(sheet, final_table)

output_path = filedialog.asksaveasfilename(title="Сохранить файл как", defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
if output_path:
    wb.save(output_path)
    print(f"Файл сохранен: {output_path}")
else:
    print("Сохранение отменено.")
