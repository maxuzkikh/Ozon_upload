import os
import openpyxl
import pyautogui
import pyperclip
import time
import signal
import sys


# Signal handler function
def signal_handler(sig, frame):
    print("Received signal. Stopping script...")
    sys.exit(0)


# Register the signal handler for SIGINT (Ctrl+C)
signal.signal(signal.SIGINT, signal_handler)


def process_image(print_path, rotate=False):
    # Simulate Ctrl+I
    pyautogui.hotkey('ctrl', 'i')
    if start == 0:
        time.sleep(3)

    time.sleep(0.1)
    # Copy the text to the clipboard
    pyperclip.copy(print_path)
    time.sleep(0.1)
    # Simulate Ctrl+V to paste the text
    pyautogui.hotkey('ctrl', 'v')
    time.sleep(0.1)
    # Press Enter to confirm
    pyautogui.press('enter')
    # Press Enter to confirm
    pyautogui.press('enter')
    time.sleep(0.1)
    # Pick Tool
    pyautogui.click(30, 100)
    pyautogui.moveTo(820, 450)

    # Rotate if specified
    if rotate:
        # Rotate Tool
        pyautogui.click(20, 189)
        # moveTo Center of Aplication
        pyautogui.moveTo(982, 555)
        # Press and hold the Shift key
        pyautogui.keyDown('shift')
        # Drag the mouse cursor downwards by 30 pixels
        pyautogui.dragRel(0, 30, duration=0.3)
        # Release the Shift key
        pyautogui.keyUp('shift')
        pyautogui.mouseUp()
        pyautogui.keyDown('ctrl')
        time.sleep(0.1)
        # Pick Tool
        pyautogui.click(30, 100)
        pyautogui.moveTo(945, 540)
        time.sleep(0.1)



    # Раскладка Copy in width!!! Window 0,100px
    layout_width = worksheet.cell(row=row, column=column_indices.get("Раскладка в ширину")).value
    if layout_width>1:
        # Press 'm' key
        time.sleep(0.1)
        pyautogui.keyDown('ctrl')
        pyautogui.press('m')
        pyautogui.keyUp('ctrl')
        # Zero values
        time.sleep(0.1)
        pyautogui.doubleClick(1020, 190)
        pyautogui.typewrite('0')
        time.sleep(0.1)
        pyautogui.doubleClick(930, 190)
        pyautogui.typewrite('100')
        time.sleep(0.1)
        # Copy Button
        pyautogui.click(1000, 360)
        time.sleep(0.1)
        # Pick Tool
        pyautogui.click(30, 100)
        pyautogui.moveTo(945, 540)
        time.sleep(0.1)
        # Select images
        pyautogui.mouseDown()
        pyautogui.dragTo(1015, 571, duration=0.3)
        # close Copy window
        pyautogui.click(1058, 366)
        time.sleep(0.1)

        # Align in width
        # Open window Align
        pyautogui.click(500, 70)
        pyautogui.doubleClick(235, 163)
        pyautogui.typewrite('4')
        time.sleep(0.1)
        # Align 4mm width
        pyautogui.click(197, 163)
        pyautogui.hotkey('ctrl', 'g')
        time.sleep(0.1)
        # Align to center of document
        pyautogui.click(284, 141)
        time.sleep(0.1)


    #time.sleep(3)
    # Select images
    #pyautogui.mouseDown()
    #pyautogui.dragTo(1015, 571, duration=0.3)
    #time.sleep(1)

    # Open window Align
    pyautogui.click(500, 70)
    time.sleep(0.1)
    # Click Center of Aplication
    pyautogui.click(982, 555)
    # Align to center of document
    pyautogui.click(284, 141)
    time.sleep(0.1)
    # Close window
    pyautogui.click(280, 111)
    time.sleep(0.1)
    # Click Center of Aplication
    pyautogui.click(982, 555)

    # Make Copies of on images from demand Num_Copies
    # Select images
    # Click Center of Aplication
    pyautogui.click(982, 555)
    # Open window Copys
    time.sleep(0.1)
    pyautogui.keyDown('ctrl')
    pyautogui.press('m')
    pyautogui.keyUp('ctrl')
    time.sleep(0.1)
    pyautogui.doubleClick(943, 185)
    pyautogui.typewrite('0')
    time.sleep(0.1)
    pyautogui.doubleClick(1024, 189)
    pyautogui.typewrite('100')

    # Get the number of copies from the corresponding cell in the "Num_Copies" column
    num_copies = worksheet.cell(row=row, column=column_indices.get("Num_Copies")).value

    # Get the value of "Раскладка в ширину" for the current row
    layout_width = worksheet.cell(row=row, column=column_indices.get("Раскладка в ширину")).value

    if num_copies is not None and isinstance(num_copies, int) and num_copies > 0:

        # Ensure num_copies is even
        if num_copies % 2 != 0:
            num_copies += 1
        # Calculate the number of copies based on "Раскладка в ширину" and round to the nearest integer
        num_copies = round(num_copies / layout_width)-1



        for _ in range(num_copies):
            pyautogui.click(1000, 364)
            # Optionally, add a small delay between clicks for stability
            time.sleep(0.05)

    time.sleep(0.05)
    # close window
    pyautogui.click(1071, 364)

    if start == 0:
        # Move images first image UP
        time.sleep(0.1)
        pyautogui.keyDown('ctrl')
        pyautogui.press('m')
        pyautogui.keyUp('ctrl')
        time.sleep(0.1)
        pyautogui.doubleClick(943, 185)
        pyautogui.typewrite('0')
        time.sleep(0.1)
        pyautogui.doubleClick(1024, 189)
        pyautogui.typewrite(str(offset))
        pyautogui.click(920, 360)
        time.sleep(0.1)
        # close window
        pyautogui.click(1071, 364)

    # Align all in height
    # select all images
    pyautogui.moveTo(943, 110)
    pyautogui.mouseDown()
    pyautogui.dragTo(1020, 995, duration=0.3)
    # Open window
    pyautogui.click(500, 70)
    pyautogui.doubleClick(223, 209)
    pyautogui.typewrite('5')
    time.sleep(0.1)
    # Align 5mm height
    pyautogui.click(196, 209)
    time.sleep(0.1)
    # Close window
    pyautogui.click(280, 111)
    time.sleep(0.1)
    # Click Center of Aplication
    pyautogui.click(982, 555)

    # Move canvas
    if start > 0:
        pyautogui.click(26, 228)
        # Click Center of Aplication
        pyautogui.moveTo(982, 555)
        # Simulate Ctrl + Left Mouse Click
        pyautogui.keyDown('ctrl')
        pyautogui.mouseDown(button='left')

        # Drag the mouse cursor upwards by 200 pixels
        pyautogui.dragRel(0, -30, duration=0.5)  # Specify the distance and duration as per your requirement
        pyautogui.keyUp('ctrl')


# Path to the file to open
file_path = r"C:\work\baby prints\MainTop\tif\Template.tpf"

# Check if file exists
if os.path.exists(file_path):
    # Open the file
    print("Opening file:", file_path)
    os.startfile(file_path)
    time.sleep(2)  # Introduce a small delay

    # Path to the Excel file
    excel_file = r"C:\Users\Max\Documents\GitHub\Ozon_upload\MainTop\WB_demand.xlsx"
    excel_file_path = r"C:\Users\Max\Documents\GitHub\Ozon_upload\barcode\Data path barcode.xlsx"

    # Check if Excel file exists
    if os.path.exists(excel_file):
        # Load the workbook
        workbook = openpyxl.load_workbook(excel_file)

        # Select the active worksheet
        worksheet = workbook.active

        # Dictionary to store column indices
        column_indices = {}

        # Iterate over the first row to find the columns
        for cell in worksheet[1]:
            value = cell.value
            if value == "Артикул":
                column_indices["Артикул"] = cell.column
            elif value == "Num_Copies":
                column_indices["Num_Copies"] = cell.column
            elif value == "путь к печати":
                column_indices["путь к печати"] = cell.column
            elif value == "Rotate":
                column_indices["Rotate"] = cell.column
            elif value == "Раскладка в ширину":  # New column
                column_indices["Раскладка в ширину"] = cell.column

        # Initialize offset to -100
        offset = -12500
        start = 0

        # Iterate over all rows in the Excel worksheet
        for row in range(2, worksheet.max_row + 1):  # Start from the second row
            # Get the value from the "Артикул" column
            article = worksheet.cell(row=row, column=column_indices.get("Артикул"))
            if article.value:
                # Get the print path from the corresponding cell in the "путь к печати" column
                print_path = worksheet.cell(row=row, column=column_indices.get("путь к печати")).value
                # Check if the print path is not empty or 0
                if print_path and print_path != 0:
                    # Check if Rotate column exists
                    if "Rotate" in column_indices:
                        # Get the value from the "Rotate" column
                        rotate_value = worksheet.cell(row=row, column=column_indices.get("Rotate")).value
                        # Convert the value to boolean
                        rotate = bool(rotate_value)
                    else:
                        rotate = False  # If Rotate column not found, set rotate to False by default
                    # Print the article, print path, and rotation values
                    # print("Article:", article.value)
                    # print("Print Path:", print_path)
                    # print("Rotate:", rotate)
                    # Call process_image function for the print path with rotation information
                    process_image(print_path, rotate=rotate)
                    offset += 200  # Decrement offset by 100 for the next iteration
                    start += 1
                else:
                    print("Skipping row:", row, "as 'путь к печати' is empty or 0")
            else:
                print("Article value not found in row:", row)

        # Close the workbook
        workbook.close()
    else:
        print("Excel file not found:", excel_file)
else:
    print("File not found:", file_path)